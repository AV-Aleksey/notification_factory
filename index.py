import subprocess
from openpyxl import load_workbook


def extract_data(startCol: int, endCol: int, startRow: int, endRow: int, filename = './kanban_tasks.xlsx'):
    workbook = load_workbook(filename=filename)

    sheet = workbook.active

    data = []

    for row in sheet.iter_rows(min_row=startRow, max_row=endRow, min_col=startCol, max_col=endCol, values_only=True):
        if row[0] is not None or row[1] is not None:
            data.append({ "title": row[0], "date": row[2].strftime("%d-%m-%Y") })
    
    return data

# важно и срочно
i_and_u = extract_data(1, 3, 5, 19)
# важно и не срочно
i_and_not_u = extract_data(6, 9, 5, 19)
# не важно и срочно
u_and_u = extract_data(1, 3, 23, 37)
# не важно и не срочно
second2 = extract_data(6, 9, 23, 37)

print()


def create_reminder(title, message, date):
    script = f'tell application "Reminders" to make new reminder with properties {{name:"{title}", body:"{message}", remind me date:date "{date}" as date}}'
    subprocess.run(["osascript", "-e", script], text=True)

# create_reminder("Важное событие", "Не забудьте выполнить задачу", "27-02-2024 23:00:00")